import streamlit as st
import pdfplumber
import re
import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────
# PARSING
# ─────────────────────────────────────────────
MONTH_ORDER = [
    "April","May","June","July","August","September",
    "October","November","December","January","February","March"
]

def pn(s):
    """Parse Indian number string → float (preserves negatives)."""
    if not s:
        return 0.0
    s = str(s).strip().replace(',', '')
    try:
        return float(s)
    except:
        return 0.0

def extract_nums(line):
    """Extract all numbers (incl negative) from a line."""
    return re.findall(r'-?[\d,]+\.?\d*', line)

def parse_gstr1_pdf(pdf_bytes):
    full_text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"

    lines = [l.strip() for l in full_text.split('\n')]

    hdr = {
        'financial_year': '', 'tax_period': '', 'gstin': '',
        'legal_name': '', 'arn': '', 'arn_date': '',
        'sign_date': '', 'authorized_signatory': '', 'designation': ''
    }
    for line in lines:
        if 'Financial year' in line and not hdr['financial_year']:
            m = re.search(r'Financial year\s+(\S+)', line)
            if m: hdr['financial_year'] = m.group(1)
        if 'Tax period' in line and not hdr['tax_period']:
            m = re.search(r'Tax period\s+(\w+)', line)
            if m: hdr['tax_period'] = m.group(1).capitalize()
        if 'GSTIN' in line and not hdr['gstin']:
            m = re.search(r'GSTIN\s+(\S+)', line)
            if m: hdr['gstin'] = m.group(1)
        if 'Legal name' in line and not hdr['legal_name']:
            m = re.search(r'Legal name of the registered person\s+(.+)', line)
            if m: hdr['legal_name'] = m.group(1).strip()
        if '(c) ARN' in line and not hdr['arn']:
            m = re.search(r'\(c\) ARN\s+(\S+)', line)
            if m: hdr['arn'] = m.group(1)
        if '(d) ARN date' in line and not hdr['arn_date']:
            m = re.search(r'\(d\) ARN date\s+(\S+)', line)
            if m: hdr['arn_date'] = m.group(1)
        if line.startswith('Date:') and not hdr['sign_date']:
            m = re.search(r'Date:\s+(\S+)', line)
            if m: hdr['sign_date'] = m.group(1)
        if 'Name of Authorized Signatory' in line and not hdr['authorized_signatory']:
            idx = lines.index(line)
            if idx + 1 < len(lines):
                hdr['authorized_signatory'] = lines[idx + 1].strip()
        if 'Designation/Status:' in line and not hdr['designation']:
            m = re.search(r'Designation/Status:\s+(.+)', line)
            if m: hdr['designation'] = m.group(1).strip()

    # ── Table data dict ──
    T = {
        '4A': [0, 'Invoice', 0,0,0,0,0],   # [no_rec, doctype, value, igst, cgst, sgst, cess]
        '4B': [0, 'Invoice', 0,0,0,0,0],
        '5':  [0, 'Invoice', 0,0,0],        # value, igst, cess
        '6A': [0, 'Invoice', 0,0,0],
        '6A_EXPWP':  [0,'Invoice',0,0,0],
        '6A_EXPWOP': [0,'Invoice',0],
        '6B': [0, 'Invoice', 0,0,0],
        '6B_SEZWP':  [0,'Invoice',0,0,0],
        '6B_SEZWOP': [0,'Invoice',0],
        '6C': [0, 'Invoice', 0,0,0,0,0],
        '7':  [0, 'Net Value', 0,0,0,0,0],
        '8':  [0,0,0,0],   # total, nil, exempted, non-gst
        # 9A rows: [no_rec, doctype, value, igst, cgst, sgst, cess] for Amended; then diff
        '9A_4A_amt':  [0,'Invoice',0,0,0,0,0],
        '9A_4A_diff': [0,0,0,0,0],
        '9A_4B_amt':  [0,'Invoice',0,0,0,0,0],
        '9A_4B_diff': [0,0,0,0,0],
        '9A_5_amt':   [0,'Invoice',0,0,0],
        '9A_5_diff':  [0,0,0],
        '9A_6A_amt':  [0,'Invoice',0,0,0],
        '9A_6A_diff': [0,0,0],
        '9A_6A_EXPWP_amt':  [0,'Invoice',0,0,0],
        '9A_6A_EXPWOP_amt': [0,'Invoice',0],
        '9A_6B_amt':  [0,'Invoice',0,0,0],
        '9A_6B_diff': [0,0,0],
        '9A_6B_SEZWP_amt':  [0,'Invoice',0,0,0],
        '9A_6B_SEZWOP_amt': [0,'Invoice',0],
        '9A_6C_amt':  [0,'Invoice',0,0,0,0,0],
        '9A_6C_diff': [0,0,0,0,0],
        # 9B CDNR
        '9B_CDNR_total':   [0,'Note',0,0,0,0,0],
        '9B_CDNR_B2B':     [0,'Note',0,0,0,0,0],
        '9B_CDNR_RC':      [0,'Note',0,0,0,0,0],
        '9B_CDNR_SEZ':     [0,'Note',0,0,0],
        '9B_CDNR_DE':      [0,'Note',0,0,0,0,0],
        # 9B CDNUR
        '9B_CDNUR_total':  [0,'Note',0,0,0],
        '9B_CDNUR_B2CL':   [0,'Note',0,0,0],
        '9B_CDNUR_EXPWP':  [0,'Note',0,0,0],
        '9B_CDNUR_EXPWOP': [0,'Note',0],
        # 9C CDNRA
        '9C_CDNRA_amt':    [0,'Note',0,0,0,0,0],
        '9C_CDNRA_diff':   [0,0,0,0,0],
        '9C_CDNRA_B2B_diff':  [0,'Note',0,0,0,0,0],
        '9C_CDNRA_RC_diff':   [0,'Note',0,0,0,0,0],
        '9C_CDNRA_SEZ_diff':  [0,'Note',0,0,0],
        '9C_CDNRA_DE_diff':   [0,'Note',0,0,0,0,0],
        # 9C CDNURA
        '9C_CDNURA_amt':   [0,'Note',0,0,0],
        '9C_CDNURA_diff':  [0,0,0],
        '9C_CDNURA_B2CL':  [0,'Note',0,0,0],
        '9C_CDNURA_EXPWP': [0,'Note',0,0,0],
        '9C_CDNURA_EXPWOP':[0,'Note',0],
        # 10
        '10_amt':  [0,'Net Value',0,0,0,0,0],
        '10_diff': [0,0,0,0,0],
        # 11A(1)
        '11A1': [0,'Net Value',0,0,0,0,0],
        # 11B(1)
        '11B1': [0,'Net Value',0,0,0,0,0],
        # 11A amend
        '11A_amt': [0,'Net Value',0,0,0,0,0],
        '11A_diff':[0,0,0,0,0],
        # 11B amend
        '11B_amt': [0,'Net Value',0,0,0,0,0],
        '11B_diff':[0,0,0,0,0],
        # 12 HSN
        '12_total': [0,0,0,0,0,0],
        '12_B2B':   [0,0,0,0,0,0],
        '12_B2C':   [0,0,0,0,0,0],
        # 13
        '13': 0,
        # 14
        '14_total':  [0,'Net Value',0,0,0,0,0],
        '14_52':     [0,'Net Value',0,0,0,0,0],
        '14_95':     [0,'Net Value',0,0,0,0,0],
        # 14A
        '14A_amt':  [0,'Net Value',0,0,0,0,0],
        '14A_diff': [0,'Net Value',0,0,0,0,0],
        '14A_52_amt': [0,'Net Value',0,0,0,0,0],
        '14A_52_diff':[0,'Net Value',0,0,0,0,0],
        '14A_95_amt': [0,'Net Value',0,0,0,0,0],
        '14A_95_diff':[0,'Net Value',0,0,0,0,0],
        # 15
        '15_total':  [0,0,0,0,0,0],
        '15_reg':    [0,0,0,0,0,0],
        '15_reg_regular':[0,0,0,0,0,0],
        '15_reg_de': [0,0,0,0,0,0],
        '15_reg_sezwp':[0,0,0,0],
        '15_reg_sezwop':[0,0],
        '15_unreg':  [0,0,0,0,0,0],
        # 15A(I)
        '15A1_amt':  [0,0,0,0,0,0],
        '15A1_diff': [0,0,0,0,0,0],
        '15A1_regular':[0,0,0,0,0,0],
        '15A1_de':   [0,0,0,0,0,0],
        '15A1_sezwp':[0,0,0,0],
        '15A1_sezwop':[0,0],
        # 15A(II)
        '15A2_amt':  [0,0,0,0,0,0],
        '15A2_diff': [0,0,0,0,0,0],
        # Total liability
        'total_liability': [0,0,0,0,0],
    }

    # State machine for parsing
    ctx = None
    sub_ctx = None

    for i, line in enumerate(lines):
        if not line or line.startswith('IP Address') or line.startswith('Description') or \
           line.startswith('No. of') or line.startswith('records') or line.startswith('Document'):
            continue

        # ─── Section detection ───
        if '4A -' in line and 'B2B Regular' in line:
            ctx = '4A'; continue
        if '4B -' in line and 'Reverse charge' in line:
            ctx = '4B'; continue
        if line.startswith('5 -') and 'B2CL' in line:
            ctx = '5'; continue
        if line.startswith('6A') and 'Exports' in line:
            ctx = '6A'; continue
        if line.startswith('6B -') and 'SEZ' in line:
            ctx = '6B'; continue
        if line.startswith('6C -') and 'Deemed' in line:
            ctx = '6C'; continue
        if line.startswith('7-') and 'B2CS' in line:
            ctx = '7'; continue
        if line.startswith('8 -') and 'Nil rated' in line:
            ctx = '8'; continue

        # 9A variants
        if '9A -' in line and 'B2B Regular' in line and 'table 4' in line:
            ctx = '9A_4A'; continue
        if '9A -' in line and 'Reverse charge' in line and 'table 4' in line:
            ctx = '9A_4B'; continue
        if '9A -' in line and 'B2CL' in line and 'table 5' in line:
            ctx = '9A_5'; continue
        if '9A -' in line and 'Export' in line and 'table 6A' in line:
            ctx = '9A_6A'; continue
        if '9A -' in line and 'SEZ' in line and 'table 6B' in line:
            ctx = '9A_6B'; continue
        if '9A -' in line and 'Deemed' in line and 'table 6C' in line:
            ctx = '9A_6C'; continue

        if '9B -' in line and 'Registered' in line and 'CDNR' in line and 'Unregistered' not in line:
            ctx = '9B_CDNR'; sub_ctx = 'total'; continue
        if '9B -' in line and 'Unregistered' in line and 'CDNUR' in line:
            ctx = '9B_CDNUR'; sub_ctx = 'total'; continue

        if '9C -' in line and 'Registered' in line and 'CDNRA' in line and 'Unregistered' not in line:
            ctx = '9C_CDNRA'; sub_ctx = 'amt'; continue
        if '9C -' in line and 'Unregistered' in line and 'CDNURA' in line:
            ctx = '9C_CDNURA'; sub_ctx = 'amt'; continue

        if line.startswith('10 -') and 'table 7' in line:
            ctx = '10'; continue
        if '11A(1), 11A(2)' in line and 'Advances received' in line:
            ctx = '11A1'; continue
        if '11B(1), 11B(2)' in line and 'Advance amount' in line:
            ctx = '11B1'; continue
        if '11A -' in line and 'Amendment' in line and 'advances received' in line:
            ctx = '11A'; continue
        if '11B -' in line and 'Amendment' in line and 'advances adjusted' in line:
            ctx = '11B'; continue

        if line.startswith('12 -') and 'HSN' in line:
            ctx = '12'; continue
        if line.startswith('13 -') and 'Documents' in line:
            ctx = '13'; continue
        if line.startswith('14 -') and 'E-Commerce' in line:
            ctx = '14'; continue
        if line.startswith('14A -') and 'E-Commerce' in line:
            ctx = '14A'; continue
        if line.startswith('15 -') and 'Supplies U/s' in line:
            ctx = '15'; sub_ctx = 'total'; continue
        if '15A (I)' in line:
            ctx = '15A1'; sub_ctx = 'amt'; continue
        if '15A (II)' in line:
            ctx = '15A2'; sub_ctx = 'amt'; continue
        if 'Total Liability' in line:
            ctx = 'total_liability'; 

        # ─── Value extraction ───
        nums = extract_nums(line)
        
        def n(idx, default=0.0):
            try: return pn(nums[idx])
            except: return default

        # Context-specific extraction
        if ctx == '4A':
            if line.startswith('Total') and 'Invoice' in line and len(nums) >= 6:
                T['4A'] = [n(0),'Invoice',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '4B':
            if line.startswith('Total') and 'Invoice' in line and len(nums) >= 6:
                T['4B'] = [n(0),'Invoice',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '5':
            if line.startswith('Total') and 'Invoice' in line and len(nums) >= 4:
                T['5'] = [n(0),'Invoice',n(1),n(2),n(3)]
        elif ctx == '6A':
            if line.startswith('Total') and 'Invoice' in line and len(nums) >= 4:
                T['6A'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif line.startswith('- EXPWP') and len(nums) >= 4:
                T['6A_EXPWP'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif line.startswith('- EXPWOP') and len(nums) >= 2:
                T['6A_EXPWOP'] = [n(0),'Invoice',n(1)]
        elif ctx == '6B':
            if line.startswith('Total') and 'Invoice' in line and len(nums) >= 4:
                T['6B'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif line.startswith('- SEZWP') and len(nums) >= 4:
                T['6B_SEZWP'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif line.startswith('- SEZWOP') and len(nums) >= 2:
                T['6B_SEZWOP'] = [n(0),'Invoice',n(1)]
        elif ctx == '6C':
            if line.startswith('Total') and 'Invoice' in line and len(nums) >= 6:
                T['6C'] = [n(0),'Invoice',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '7':
            if line.startswith('Total') and 'Net Value' in line and len(nums) >= 6:
                T['7'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '8':
            if line.startswith('Total') and len(nums) >= 1:
                T['8'][0] = n(0)
            elif line.startswith('- Nil') and len(nums) >= 1:
                T['8'][1] = n(0)
            elif line.startswith('- Exempted') and len(nums) >= 1:
                T['8'][2] = n(0)
            elif line.startswith('- Non-GST') and len(nums) >= 1:
                T['8'][3] = n(0)

        elif ctx == '9A_4A':
            if 'Amended amount' in line and len(nums) >= 6:
                T['9A_4A_amt'] = [n(0),'Invoice',n(1),n(2),n(3),n(4),n(5)]
            elif 'Net differential' in line and len(nums) >= 5:
                T['9A_4A_diff'] = [n(0),n(1),n(2),n(3),n(4)]
        elif ctx == '9A_4B':
            if 'Amended amount' in line and len(nums) >= 6:
                T['9A_4B_amt'] = [n(0),'Invoice',n(1),n(2),n(3),n(4),n(5)]
            elif 'Net differential' in line and len(nums) >= 5:
                T['9A_4B_diff'] = [n(0),n(1),n(2),n(3),n(4)]
        elif ctx == '9A_5':
            if 'Amended amount' in line and len(nums) >= 4:
                T['9A_5_amt'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif 'Net differential' in line and len(nums) >= 3:
                T['9A_5_diff'] = [n(0),n(1),n(2)]
        elif ctx == '9A_6A':
            if 'Amended amount' in line and len(nums) >= 4:
                T['9A_6A_amt'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif 'Net differential' in line and '- Total' in line and len(nums) >= 3:
                T['9A_6A_diff'] = [n(0),n(1),n(2)]
            elif '- EXPWP' in line and len(nums) >= 4:
                T['9A_6A_EXPWP_amt'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif '- EXPWOP' in line and len(nums) >= 2:
                T['9A_6A_EXPWOP_amt'] = [n(0),'Invoice',n(1)]
        elif ctx == '9A_6B':
            if 'Amended amount' in line and len(nums) >= 4:
                T['9A_6B_amt'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif 'Net differential' in line and '- Total' in line and len(nums) >= 3:
                T['9A_6B_diff'] = [n(0),n(1),n(2)]
            elif '- SEZWP' in line and len(nums) >= 4:
                T['9A_6B_SEZWP_amt'] = [n(0),'Invoice',n(1),n(2),n(3)]
            elif '- SEZWOP' in line and len(nums) >= 2:
                T['9A_6B_SEZWOP_amt'] = [n(0),'Invoice',n(1)]
        elif ctx == '9A_6C':
            if 'Amended amount' in line and len(nums) >= 6:
                T['9A_6C_amt'] = [n(0),'Invoice',n(1),n(2),n(3),n(4),n(5)]
            elif 'Net differential' in line and len(nums) >= 5:
                T['9A_6C_diff'] = [n(0),n(1),n(2),n(3),n(4)]

        elif ctx == '9B_CDNR':
            # Structure: Total-Net-off line, then sub-section headers + Net Total lines in order:
            # B2B Regular → Reverse charge → SEZWP/SEZWOP → DE
            if ('Net off' in line or 'Total -' in line) and 'Net Total' not in line:
                if len(nums) >= 7:
                    T['9B_CDNR_total'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]
                sub_ctx = 'await_B2B'
            elif 'B2B Regular' in line:
                sub_ctx = 'await_B2B'
            elif 'Reverse charge' in line and 'table 4' in line:
                sub_ctx = 'await_RC'
            elif 'SEZWP' in line and 'table 6B' in line:
                sub_ctx = 'await_SEZ'
            elif 'table 6C' in line:
                sub_ctx = 'await_DE'
            elif 'Net Total' in line:
                if sub_ctx == 'await_B2B' and len(nums) >= 6:
                    T['9B_CDNR_B2B'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]
                    sub_ctx = 'await_RC'
                elif sub_ctx == 'await_RC' and len(nums) >= 6:
                    T['9B_CDNR_RC'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]
                    sub_ctx = 'await_SEZ'
                elif sub_ctx == 'await_SEZ' and len(nums) >= 3:
                    T['9B_CDNR_SEZ'] = [n(0),'Note',n(1),n(2),n(3) if len(nums)>3 else 0]
                    sub_ctx = 'await_DE'
                elif sub_ctx == 'await_DE' and len(nums) >= 6:
                    T['9B_CDNR_DE'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '9B_CDNUR':
            if 'Net off' in line or 'Total -' in line:
                if len(nums) >= 4:
                    T['9B_CDNUR_total'] = [n(0),'Note',n(1),n(2),n(3)]
            elif '- B2CL' in line and len(nums) >= 4:
                T['9B_CDNUR_B2CL'] = [n(0),'Note',n(1),n(2),n(3)]
            elif '- EXPWP' in line and len(nums) >= 4:
                T['9B_CDNUR_EXPWP'] = [n(0),'Note',n(1),n(2),n(3)]
            elif '- EXPWOP' in line and len(nums) >= 2:
                T['9B_CDNUR_EXPWOP'] = [n(0),'Note',n(1)]

        elif ctx == '9C_CDNRA':
            if 'Amended amount' in line and '- Total' in line and len(nums) >= 6:
                T['9C_CDNRA_amt'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]
            elif 'Net Differential' in line and '- Total' in line and len(nums) >= 5:
                T['9C_CDNRA_diff'] = [n(0),n(1),n(2),n(3),n(4)]
            elif 'B2B Regular' in line:
                sub_ctx = 'B2B'
            elif 'Reverse charge' in line:
                sub_ctx = 'RC'
            elif 'SEZWP' in line:
                sub_ctx = 'SEZ'
            elif 'table 6C' in line:
                sub_ctx = 'DE'
            elif 'Net total' in line:
                if sub_ctx == 'B2B' and len(nums) >= 7:
                    T['9C_CDNRA_B2B_diff'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]
                elif sub_ctx == 'RC' and len(nums) >= 7:
                    T['9C_CDNRA_RC_diff'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]
                elif sub_ctx == 'SEZ' and len(nums) >= 4:
                    T['9C_CDNRA_SEZ_diff'] = [n(0),'Note',n(1),n(2),n(3)]
                elif sub_ctx == 'DE' and len(nums) >= 7:
                    T['9C_CDNRA_DE_diff'] = [n(0),'Note',n(1),n(2),n(3),n(4),n(5)]

        elif ctx == '9C_CDNURA':
            if 'Amended amount' in line and '- Total' in line and len(nums) >= 4:
                T['9C_CDNURA_amt'] = [n(0),'Note',n(1),n(2),n(3)]
            elif 'Net Differential' in line and '- Total' in line and len(nums) >= 3:
                T['9C_CDNURA_diff'] = [n(0),n(1),n(2)]
            elif '- B2CL' in line and len(nums) >= 4:
                T['9C_CDNURA_B2CL'] = [n(0),'Note',n(1),n(2),n(3)]
            elif '- EXPWP' in line and len(nums) >= 4:
                T['9C_CDNURA_EXPWP'] = [n(0),'Note',n(1),n(2),n(3)]
            elif '- EXPWOP' in line and len(nums) >= 2:
                T['9C_CDNURA_EXPWOP'] = [n(0),'Note',n(1)]

        elif ctx == '10':
            if 'Amended amount' in line and len(nums) >= 6:
                T['10_amt'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
            elif 'Net differential' in line and len(nums) >= 5:
                T['10_diff'] = [n(0),n(1),n(2),n(3),n(4)]
            elif line.startswith('Net differential') and len(nums) >= 5:
                T['10_diff'] = [n(0),n(1),n(2),n(3),n(4)]
        elif ctx == '11A1':
            if line.startswith('Total') and len(nums) >= 6:
                T['11A1'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '11B1':
            if line.startswith('Total') and len(nums) >= 6:
                T['11B1'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '11A':
            if 'Amended amount' in line and len(nums) >= 6:
                T['11A_amt'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
            elif line.startswith('Total') and len(nums) >= 5:
                T['11A_diff'] = [n(0),n(1),n(2),n(3),n(4)]
        elif ctx == '11B':
            if 'Amended amount' in line and len(nums) >= 6:
                T['11B_amt'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
            elif line.startswith('Total') and len(nums) >= 5:
                T['11B_diff'] = [n(0),n(1),n(2),n(3),n(4)]

        elif ctx == '12':
            if line.startswith('Total') and len(nums) >= 6:
                T['12_total'] = [n(0),'NA',n(1),n(2),n(3),n(4),n(5)]
            elif 'B2B Total' in line and len(nums) >= 6:
                T['12_B2B'] = [n(0),'NA',n(1),n(2),n(3),n(4),n(5)]
            elif 'B2C Total' in line and len(nums) >= 6:
                T['12_B2C'] = [n(0),'NA',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '13':
            if 'Net issued' in line and len(nums) >= 1:
                T['13'] = n(0)

        elif ctx == '14':
            if line.startswith('Total') and 'Net Value' in line and len(nums) >= 6:
                T['14_total'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
            elif '(a)' in line and '52' in line and len(nums) >= 6:
                T['14_52'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
            elif '(b)' in line and '9(5)' in line and len(nums) >= 6:
                T['14_95'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '14A':
            if 'Amended amount' in line and '– Total' in line and len(nums) >= 6:
                T['14A_amt'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]
            elif 'Net differential' in line and 'Original' in line and len(nums) >= 7:
                T['14A_diff'] = [n(0),'Net Value',n(1),n(2),n(3),n(4),n(5)]

        elif ctx == '15':
            if line.startswith('Total') and len(nums) >= 6:
                T['15_total'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif '- For Registered' in line and len(nums) >= 6:
                T['15_reg'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif '- Regular' in line and len(nums) >= 6:
                T['15_reg_regular'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif '- DE' in line and len(nums) >= 6:
                T['15_reg_de'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif '- SEZWP' in line and len(nums) >= 4:
                T['15_reg_sezwp'] = [n(0),n(1),n(2),n(3)]
            elif '- SEZWOP' in line and len(nums) >= 2:
                T['15_reg_sezwop'] = [n(0),n(1)]
            elif '- For Unregistered' in line and len(nums) >= 6:
                T['15_unreg'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
        elif ctx == '15A1':
            if 'Amended amount' in line and len(nums) >= 6:
                T['15A1_amt'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif 'Net differential' in line and len(nums) >= 6:
                T['15A1_diff'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif '- Regular' in line and len(nums) >= 6:
                T['15A1_regular'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif '- DE' in line and len(nums) >= 6:
                T['15A1_de'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif '- SEZWP' in line and len(nums) >= 4:
                T['15A1_sezwp'] = [n(0),n(1),n(2),n(3)]
            elif '- SEZWOP' in line and len(nums) >= 2:
                T['15A1_sezwop'] = [n(0),n(1)]
        elif ctx == '15A2':
            if 'Amended amount' in line and len(nums) >= 6:
                T['15A2_amt'] = [n(0),n(1),n(2),n(3),n(4),n(5)]
            elif 'Net differential' in line and len(nums) >= 6:
                T['15A2_diff'] = [n(0),n(1),n(2),n(3),n(4),n(5)]

        elif ctx == 'total_liability':
            if 'Total Liability' in line and len(nums) >= 5:
                T['total_liability'] = [n(0),n(1),n(2),n(3),n(4)]

    return hdr, T


# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────

# Colors
C_HEADER_DARK  = "1F3864"   # dark navy
C_HEADER_MID   = "2E75B6"   # medium blue
C_HEADER_LIGHT = "D6E4F0"   # light blue
C_TABLE_HEADER = "BDD7EE"
C_SUB_HEADER   = "E2EFDA"   # light green
C_TITLE_BG     = "FFF2CC"   # yellow
C_TOTAL_BG     = "FCE4D6"   # light orange
C_NEG_FONT     = "C00000"   # red for negatives
C_WHITE        = "FFFFFF"
C_BORDER_CLR   = "9DC3E6"

def mk_font(bold=False, color="000000", size=10, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def mk_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def mk_border(style="thin", color=C_BORDER_CLR):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def mk_align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_cell(cell, value=None, bold=False, bg=None, font_color="000000",
               h_align="center", v_align="center", wrap=True, size=10,
               border=True, num_fmt=None):
    if value is not None:
        cell.value = value
    cell.font = mk_font(bold=bold, color=font_color, size=size)
    if bg:
        cell.fill = mk_fill(bg)
    cell.alignment = mk_align(h_align, v_align, wrap)
    if border:
        cell.border = mk_border()
    if num_fmt:
        cell.number_format = num_fmt

def num_cell(cell, value, bold=False, bg=None):
    """Write a numeric value; red if negative."""
    fc = C_NEG_FONT if (isinstance(value, (int, float)) and value < 0) else "000000"
    num_fmt = '#,##0.00'
    style_cell(cell, value, bold=bold, bg=bg, font_color=fc, num_fmt=num_fmt)

def hdr_row(ws, row, cols_vals, bg, bold=True, font_color=C_WHITE, size=10):
    for col, val in enumerate(cols_vals, 1):
        c = ws.cell(row=row, column=col)
        style_cell(c, val, bold=bold, bg=bg, font_color=font_color, size=size)


def _sum_total_row(ws, row, start_col, end_col, data_start, data_end, bg=None):
    """Write SUM formula cells for a total row."""
    bg = bg or C_TOTAL_BG
    for col in range(start_col, end_col + 1):
        cl = get_column_letter(col)
        c = ws.cell(row=row, column=col)
        c.value = f"=SUM({cl}{data_start}:{cl}{data_end})"
        c.font = mk_font(bold=True)
        c.fill = mk_fill(bg)
        c.alignment = mk_align()
        c.border = mk_border()
        c.number_format = '#,##0.00'



def build_excel(all_data):
    """Build full Excel workbook from list of (hdr, T) dicts."""
    wb = Workbook()

    # ─── Sheet 1: SUMMARY ───
    ws_sum = wb.active
    ws_sum.title = "SUMMARY"
    _build_summary(ws_sum, all_data)

    # ─── Sheet 2: sales,exports,sez ───
    ws_sales = wb.create_sheet("sales,exports,sez")
    _build_sales(ws_sales, all_data)

    # ─── Sheet 3: CDNR ───
    ws_cdnr = wb.create_sheet("CDNR")
    _build_cdnr(ws_cdnr, all_data)

    # ─── Sheet 4: Remaining ───
    ws_rem = wb.create_sheet("Remaining")
    _build_remaining(ws_rem, all_data)

    # ─── Sheet 5: all data ───
    ws_all = wb.create_sheet("all data")
    _build_all_data(ws_all, all_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _merge_title(ws, row, start_col, end_col, text, bg=C_HEADER_DARK, font_color=C_WHITE, bold=True, size=11):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col)
    style_cell(c, text, bold=bold, bg=bg, font_color=font_color, size=size, h_align="center")


def _build_summary(ws, all_data):
    """Build SUMMARY sheet matching sample_file.xlsx exactly:
    - Rows 1-15:  Filing info table (FY merged, months, GSTIN, ARN)
    - Row 16:     Section title
    - Row 17:     Group headers (merged)
    - Row 18:     Field headers
    - Rows 19-30: APRIL-MARCH with exact cross-sheet formulas
    - Row 31:     SUM totals
    - Rows 33+:   FILING DETAILS table
    """
    from collections import defaultdict

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    for col in 'EFGHIJKLMNOPQ':
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['R'].width = 38

    # ── ROW 1: Main title ──
    ws.merge_cells('A1:R1')
    c = ws['A1']
    style_cell(c, 'GSTR-1 CONSOLIDATED SUMMARY', bold=True,
               bg=C_HEADER_DARK, font_color=C_WHITE, size=14)

    # ── ROW 2: Filing table headers ──
    for col, hdr in enumerate(['Financial Year','Tax Period','GSTIN','ARN DATE'], 1):
        style_cell(ws.cell(row=2, column=col), hdr,
                   bold=True, bg=C_HEADER_MID, font_color=C_WHITE, size=10)

    # ── ROWS 3-14: Filing data (APRIL=row3 … MARCH=row14) ──
    # Collect all PDFs, group by (FY, GSTIN), sort months
    # Build a month→hdr map for the 12 fixed rows
    month_hdr = {}   # month_name → hdr dict (last one wins if duplicates)
    for d in all_data:
        month_hdr[d['hdr']['tax_period']] = d['hdr']

    # FY for merge — use first found
    fy = all_data[0]['hdr']['financial_year'] if all_data else '2025-26'
    fy_start_row = 3

    for i, mon in enumerate(MONTH_ORDER):
        r = 3 + i   # APRIL→row3, MAY→row4 … MARCH→row14
        alt_bg = C_HEADER_LIGHT if r % 2 == 0 else C_WHITE
        h = month_hdr.get(mon)
        gstin   = h['gstin']   if h else ''
        arn_dt  = h['arn_date'] if h else ''
        for col, val in enumerate([fy, mon.upper(), gstin, arn_dt], 1):
            c = ws.cell(row=r, column=col)
            style_cell(c, val, h_align='center', bg=alt_bg)

    # Merge FY column A rows 3-14
    ws.merge_cells(start_row=3, start_column=1, end_row=14, end_column=1)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    ws['A3'].font      = mk_font(bold=True, size=10)

    # ── ROW 16: Reconciliation section title ──
    ws.merge_cells('B16:R16')
    style_cell(ws['B16'], 'GSTR-1 MONTHLY RECONCILIATION SUMMARY',
               bold=True, bg=C_HEADER_DARK, font_color=C_WHITE, size=12)

    # ── ROW 17: Group headers ──
    ws.merge_cells('C17:G17')
    style_cell(ws['C17'],
               'summary-B2B, SEZ, EXP, B2CS (excl Amendments)',
               bold=True, bg=C_HEADER_MID, font_color=C_WHITE, size=10)
    ws.merge_cells('H17:L17')
    style_cell(ws['H17'],
               'Total Liability (Outward supplies other than Reverse charge)',
               bold=True, bg=C_HEADER_MID, font_color=C_WHITE, size=10)
    ws.merge_cells('M17:Q17')
    style_cell(ws['M17'], 'Diff',
               bold=True, bg='E2EFDA', font_color='000000', size=10)
    style_cell(ws['R17'], 'Remarks',
               bold=True, bg=C_TABLE_HEADER, font_color='000000', size=10)
    style_cell(ws['B17'], '',
               bold=True, bg=C_HEADER_DARK, font_color=C_WHITE, size=10)

    # ── ROW 18: Field headers ──
    fld_hdrs = ['Month',
                'Taxable Value','Integrated Tax','Central Tax','State Tax','Cess',
                'Taxable Value','Integrated Tax','Central Tax','State Tax','Cess',
                'Taxable Value','Integrated Tax','Central Tax','State Tax','Cess',
                'Remarks']
    for col, lbl in enumerate(fld_hdrs, 2):   # col B=2 … R=18
        style_cell(ws.cell(row=18, column=col), lbl,
                   bold=True, bg=C_TABLE_HEADER, font_color='000000', size=9)

    # ── ROWS 19-30: APRIL(row19)…MARCH(row30) with exact cross-sheet formulas ──
    # sales sheet data rows: APRIL=row4, MAY=row5 … MARCH=row15
    # Remaining Total Liability: APRIL=row342, MAY=row343 … MARCH=row353
    SLS = "'sales,exports,sez'"
    REM = 'Remaining'
    CDN = 'CDNR'

    DATA_ROW_START = 19   # APRIL
    DATA_ROW_END   = 30   # MARCH

    for i, mon in enumerate(MONTH_ORDER):
        r   = DATA_ROW_START + i   # summary row (19..30)
        sr  = 4 + i                # sales/CDNR sheet row (4..15)
        rr  = 342 + i              # Remaining Total Liability row (342..353)
        alt_bg = C_WHITE if r % 2 == 0 else 'F5FBFF'

        # Col B: Month name
        style_cell(ws.cell(row=r, column=2), mon.upper(),
                   bold=True, bg=alt_bg, h_align='center')

        def wf(col, val, diff=False, remark=False, zero=False):
            """Write formula or value to a summary cell."""
            c = ws.cell(row=r, column=col)
            c.value = val
            fc = C_NEG_FONT if diff else '000000'
            c.font  = mk_font(bold=remark, color=fc, size=10)
            c.fill  = mk_fill('FFF2CC' if diff else alt_bg)
            c.alignment = mk_align(h='left' if remark else 'center')
            c.border = mk_border()
            if not remark:
                c.number_format = '#,##0.00'

        # ── C: Taxable Value (B2B+EXP+SEZ+6C+B2CS+Nil+Exempted+NonGST + CDNR B2B + CDNR SEZ) ──
        wf(3, (f"={SLS}!C{sr}+{SLS}!P{sr}+{SLS}!S{sr}+{SLS}!T{sr}"
               f"+{SLS}!W{sr}+{SLS}!AC{sr}+{SLS}!AH{sr}+{SLS}!AI{sr}"
               f"+{SLS}!AJ{sr}+{CDN}!C{sr}+{CDN}!M{sr}"))

        # ── D: Integrated Tax ──
        wf(4, (f"={SLS}!D{sr}+{SLS}!Q{sr}+{SLS}!U{sr}"
               f"+{CDN}!D{sr}+{CDN}!N{sr}+{SLS}!AD{sr}"))

        # ── E: Central Tax ──
        wf(5, f"={SLS}!E{sr}+{SLS}!AE{sr}+{CDN}!E{sr}+{CDN}!O{sr}")

        # ── F: State Tax ──
        wf(6, f"={SLS}!F{sr}+{SLS}!AF{sr}+{CDN}!F{sr}+{CDN}!P{sr}")

        # ── G: Cess (hardcoded 0 as per sample) ──
        wf(7, 0)

        # ── H-L: Total Liability from Remaining sheet ──
        for col, rcol in zip(range(8, 13), ['C','D','E','F','G']):
            wf(col, f"={REM}!{rcol}{rr}")

        # ── M-Q: Diff = Total Liability - B2B Summary ──
        diff_map = {13:('H','C'), 14:('I','D'), 15:('J','E'),
                    16:('K','F'), 17:('L','G')}
        for dcol, (tc, bc) in diff_map.items():
            wf(dcol, f"={tc}{r}-{bc}{r}", diff=True)

        # ── R: Remarks IF formula ──
        wf(18,
           f'=IF((ABS(M{r})+ABS(N{r})+ABS(O{r})+ABS(P{r})+ABS(Q{r}))=0,'
           f'"Okay","Need to check other tables or Amendments")',
           remark=True)

    # ── ROW 31: Total row with SUM(C19:C30) formulas ──
    style_cell(ws.cell(row=31, column=2), 'Total',
               bold=True, bg=C_TOTAL_BG, h_align='center')
    for col in range(3, 18):
        cl = get_column_letter(col)
        c  = ws.cell(row=31, column=col)
        c.value          = f'=SUM({cl}{DATA_ROW_START}:{cl}{DATA_ROW_END})'
        c.font           = mk_font(bold=True)
        c.fill           = mk_fill(C_TOTAL_BG)
        c.alignment      = mk_align()
        c.border         = mk_border()
        c.number_format  = '#,##0.00'
    style_cell(ws.cell(row=31, column=18), '', bold=True, bg=C_TOTAL_BG)

    # ── ROWS 33+: FILING DETAILS ──
    row = 33
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=4)
    style_cell(ws.cell(row=row, column=1), 'FILING DETAILS',
               bold=True, bg=C_HEADER_MID, font_color=C_WHITE, size=11)
    row += 1
    for col, hdr in enumerate(['GSTIN','Legal Name','ARN','ARN Date'], 1):
        style_cell(ws.cell(row=row, column=col), hdr,
                   bold=True, bg=C_TABLE_HEADER, font_color='000000')
    row += 1
    seen = set()
    for d in all_data:
        h = d['hdr']
        key = (h['gstin'], h['tax_period'])
        if key not in seen:
            seen.add(key)
            for col, val in enumerate(
                    [h['gstin'], h['legal_name'], h['arn'], h['arn_date']], 1):
                style_cell(ws.cell(row=row, column=col), val,
                           bg=C_WHITE,
                           h_align='left' if col == 2 else 'center')
            row += 1


def _build_sales(ws, all_data):
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 14
    for col in range(3, 50):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Row 1: table group headers
    row = 1
    headers_groups = [
        ("", 1, 2),
        ("4A - B2B Regular", 3, 7),
        ("4B - B2B Reverse Charge", 8, 12),
        ("5 - B2CL (Large)", 13, 15),
        ("6A - Exports (EXPWP / EXPWOP)", 16, 19),
        ("6B - SEZ (SEZWP / SEZWOP)", 20, 23),
        ("6C - Deemed Exports", 24, 28),
        ("7 - B2CS (Others)", 29, 33),
        ("8 - Nil / Exempted / Non-GST", 34, 36),
    ]
    for label, sc, ec in headers_groups:
        if ec > sc:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        c = ws.cell(row=row, column=sc)
        style_cell(c, label, bold=True, bg=C_HEADER_DARK, font_color=C_WHITE, size=10)

    # Row 2: sub-group headers
    row = 2
    sub_groups = [
        ("EXPWP", 16, 18), ("EXPWOP", 19, 19),
        ("SEZWP", 20, 22), ("SEZWOP", 23, 23),
    ]
    for sg_label, sc, ec in sub_groups:
        if ec > sc:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        c = ws.cell(row=row, column=sc)
        style_cell(c, sg_label, bold=True, bg=C_HEADER_MID, font_color=C_WHITE)
    from openpyxl.cell.cell import MergedCell
    for col in range(1, 37):
        c = ws.cell(row=row, column=col)
        if not isinstance(c, MergedCell) and not c.value:
            style_cell(c, "", bg=C_HEADER_MID, font_color=C_WHITE)

    # Row 3: field labels
    row = 3
    fields = [
        "", "Month",
        "Taxable Value", "IGST", "CGST", "SGST", "Cess",   # 4A
        "Taxable Value", "IGST", "CGST", "SGST", "Cess",   # 4B
        "Taxable Value", "IGST", "Cess",                    # 5
        "Taxable Value", "IGST", "Cess",                    # 6A EXPWP
        "Taxable Value",                                     # 6A EXPWOP
        "Taxable Value", "IGST", "Cess",                    # 6B SEZWP
        "Taxable Value",                                     # 6B SEZWOP
        "Taxable Value", "IGST", "CGST", "SGST", "Cess",   # 6C
        "Taxable Value", "IGST", "CGST", "SGST", "Cess",   # 7
        "Nil", "Exempted", "Non-GST",                       # 8
    ]
    for col, lbl in enumerate(fields, 1):
        c = ws.cell(row=row, column=col)
        style_cell(c, lbl, bold=True, bg=C_TABLE_HEADER, font_color="000000", size=9)

    # Build months map
    months_seen = {}
    for d in all_data:
        mon = d['hdr']['tax_period']
        if mon not in months_seen:
            months_seen[mon] = []
        months_seen[mon].append((d['hdr'], d['T']))

    row = 4
    data_start = row

    for mon in MONTH_ORDER:
        bg = C_WHITE if row % 2 == 0 else "F5FBFF"
        ws.cell(row=row, column=1).value = ""
        style_cell(ws.cell(row=row, column=2), mon.upper(), bold=True,
                   bg=C_HEADER_LIGHT if mon not in months_seen else bg,
                   h_align="center")
        if mon not in months_seen:
            for col in range(3, 37):
                num_cell(ws.cell(row=row, column=col), 0.0, bg=C_HEADER_LIGHT)
            row += 1
            continue
        for h, T in months_seen[mon]:
            vals = [
                T['4A'][2], T['4A'][3], T['4A'][4], T['4A'][5], T['4A'][6],
                T['4B'][2], T['4B'][3], T['4B'][4], T['4B'][5], T['4B'][6],
                T['5'][2],  T['5'][3],  T['5'][4]  if len(T['5'])>4  else 0,
                T['6A_EXPWP'][2], T['6A_EXPWP'][3],
                    T['6A_EXPWP'][4] if len(T['6A_EXPWP'])>4 else 0,
                T['6A_EXPWOP'][2] if len(T['6A_EXPWOP'])>2 else 0,
                T['6B_SEZWP'][2], T['6B_SEZWP'][3],
                    T['6B_SEZWP'][4] if len(T['6B_SEZWP'])>4 else 0,
                T['6B_SEZWOP'][2] if len(T['6B_SEZWOP'])>2 else 0,
                T['6C'][2], T['6C'][3], T['6C'][4], T['6C'][5], T['6C'][6],
                T['7'][2],  T['7'][3],  T['7'][4],  T['7'][5],  T['7'][6],
                T['8'][1],  T['8'][2],  T['8'][3],
            ]
            for ci, v in enumerate(vals, 3):
                num_cell(ws.cell(row=row, column=ci), v, bg=bg)
        row += 1

    data_end = row - 1
    # Total row — SUM formulas
    ws.cell(row=row, column=1).value = ""
    style_cell(ws.cell(row=row, column=2), "TOTAL", bold=True,
               bg=C_TOTAL_BG, h_align="center")
    _sum_total_row(ws, row, 3, 36, data_start, data_end)


def _build_cdnr(ws, all_data):
    for col in range(1, 25):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 14

    # ── CDNR Registered ──
    _merge_title(ws, 1, 1, 22,
                 "9B - Credit/Debit Notes (Registered) – CDNR", C_HEADER_DARK)
    row = 2
    groups = [
        ("", 1, 2), ("B2B Regular", 3, 7), ("B2B Reverse Charge", 8, 12),
        ("SEZWP/SEZWOP", 13, 17), ("DE", 18, 22),
    ]
    for label, sc, ec in groups:
        if ec > sc:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        style_cell(ws.cell(row=row, column=sc), label,
                   bold=True, bg=C_HEADER_MID, font_color=C_WHITE)
    row = 3
    for col, lbl in enumerate(
            ["", "Month"] + ["Taxable Value","IGST","CGST","SGST","Cess"]*4, 1):
        style_cell(ws.cell(row=row, column=col), lbl,
                   bold=True, bg=C_TABLE_HEADER, font_color="000000", size=9)

    months_seen = {}
    for d in all_data:
        mon = d['hdr']['tax_period']
        if mon not in months_seen:
            months_seen[mon] = []
        months_seen[mon].append((d['hdr'], d['T']))

    row = 4
    data_start = row
    for mon in MONTH_ORDER:
        bg = C_WHITE if row % 2 == 0 else "F5FBFF"
        style_cell(ws.cell(row=row, column=2), mon.upper(), bold=True,
                   bg=C_HEADER_LIGHT if mon not in months_seen else bg,
                   h_align="center")
        if mon not in months_seen:
            for col in range(3, 23):
                num_cell(ws.cell(row=row, column=col), 0.0, bg=C_HEADER_LIGHT)
            row += 1
            continue
        for h, T in months_seen[mon]:
            vals = [
                T['9B_CDNR_B2B'][2], T['9B_CDNR_B2B'][3],
                T['9B_CDNR_B2B'][4], T['9B_CDNR_B2B'][5], T['9B_CDNR_B2B'][6],
                T['9B_CDNR_RC'][2],  T['9B_CDNR_RC'][3],
                T['9B_CDNR_RC'][4],  T['9B_CDNR_RC'][5],  T['9B_CDNR_RC'][6],
                T['9B_CDNR_SEZ'][2], T['9B_CDNR_SEZ'][3],
                T['9B_CDNR_SEZ'][4] if len(T['9B_CDNR_SEZ'])>4 else 0, 0, 0,
                T['9B_CDNR_DE'][2],  T['9B_CDNR_DE'][3],
                T['9B_CDNR_DE'][4],  T['9B_CDNR_DE'][5],  T['9B_CDNR_DE'][6],
            ]
            for ci, v in enumerate(vals, 3):
                num_cell(ws.cell(row=row, column=ci), v, bg=bg)
        row += 1
    data_end = row - 1

    style_cell(ws.cell(row=row, column=2), "TOTAL", bold=True,
               bg=C_TOTAL_BG, h_align="center")
    _sum_total_row(ws, row, 3, 22, data_start, data_end)

    # ── CDNUR Unregistered ──
    row += 2
    _merge_title(ws, row, 1, 9,
                 "9B - Credit/Debit Notes (Unregistered) – CDNUR", C_HEADER_DARK)
    row += 1
    cdnur_grps = [("",1,2),("B2CL",3,5),("EXPWP",6,8),("EXPWOP",9,9)]
    for label, sc, ec in cdnur_grps:
        if ec > sc:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        style_cell(ws.cell(row=row, column=sc), label,
                   bold=True, bg=C_HEADER_MID, font_color=C_WHITE)
    row += 1
    for col, lbl in enumerate(
            ["","Month","Taxable Value","IGST","Cess",
             "Taxable Value","IGST","Cess","Taxable Value"], 1):
        style_cell(ws.cell(row=row, column=col), lbl,
                   bold=True, bg=C_TABLE_HEADER, font_color="000000", size=9)
    row += 1
    data_start2 = row
    for mon in MONTH_ORDER:
        bg = C_WHITE if row % 2 == 0 else "F5FBFF"
        style_cell(ws.cell(row=row, column=2), mon.upper(), bold=True,
                   bg=C_HEADER_LIGHT if mon not in months_seen else bg,
                   h_align="center")
        if mon not in months_seen:
            for col in range(3, 10):
                num_cell(ws.cell(row=row, column=col), 0.0, bg=C_HEADER_LIGHT)
            row += 1
            continue
        for h, T in months_seen[mon]:
            vals = [
                T['9B_CDNUR_B2CL'][2],  T['9B_CDNUR_B2CL'][3],
                T['9B_CDNUR_B2CL'][4]  if len(T['9B_CDNUR_B2CL'])>4  else 0,
                T['9B_CDNUR_EXPWP'][2], T['9B_CDNUR_EXPWP'][3],
                T['9B_CDNUR_EXPWP'][4] if len(T['9B_CDNUR_EXPWP'])>4 else 0,
                T['9B_CDNUR_EXPWOP'][2] if len(T['9B_CDNUR_EXPWOP'])>2 else 0,
            ]
            for ci, v in enumerate(vals, 3):
                num_cell(ws.cell(row=row, column=ci), v, bg=bg)
        row += 1
    data_end2 = row - 1

    style_cell(ws.cell(row=row, column=2), "TOTAL", bold=True,
               bg=C_TOTAL_BG, h_align="center")
    _sum_total_row(ws, row, 3, 9, data_start2, data_end2)


def _build_remaining(ws, all_data):
    """Tables 9A to 15A(II) + Total Liability – all Total rows use SUM formulas."""
    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22

    months_seen = {}
    for d in all_data:
        mon = d['hdr']['tax_period']
        if mon not in months_seen:
            months_seen[mon] = []
        months_seen[mon].append((d['hdr'], d['T']))

    row = 1

    def sec_title(title, ncols):
        nonlocal row
        _merge_title(ws, row, 1, ncols, title, C_HEADER_DARK)
        row += 1

    def sub_hdr(groups):
        nonlocal row
        for label, sc, ec in groups:
            if ec > sc:
                ws.merge_cells(start_row=row, start_column=sc,
                               end_row=row, end_column=ec)
            style_cell(ws.cell(row=row, column=sc), label,
                       bold=True, bg=C_HEADER_MID, font_color=C_WHITE)
        row += 1

    def fld_hdr(fields):
        nonlocal row
        for col, lbl in enumerate(fields, 1):
            style_cell(ws.cell(row=row, column=col), lbl,
                       bold=True, bg=C_TABLE_HEADER, font_color="000000", size=9)
        row += 1

    def data_rows(get_vals, n_val_cols, start_col=3):
        """Write data rows for all months + a SUM-formula Total row.
        Returns the row after the Total row."""
        nonlocal row
        data_start = row
        for mon in MONTH_ORDER:
            bg = C_WHITE if row % 2 == 0 else "F5FBFF"
            style_cell(ws.cell(row=row, column=2), mon.upper(), bold=True,
                       bg=C_HEADER_LIGHT if mon not in months_seen else bg,
                       h_align="center")
            if mon not in months_seen:
                for col in range(start_col, start_col + n_val_cols):
                    num_cell(ws.cell(row=row, column=col), 0.0, bg=C_HEADER_LIGHT)
            else:
                for h, T in months_seen[mon]:
                    vals = get_vals(h, T)
                    for ci, v in enumerate(vals, start_col):
                        num_cell(ws.cell(row=row, column=ci), v, bg=bg)
            row += 1
        data_end = row - 1
        # Total row with SUM formulas
        style_cell(ws.cell(row=row, column=2), "TOTAL", bold=True,
                   bg=C_TOTAL_BG, h_align="center")
        _sum_total_row(ws, row, start_col, start_col + n_val_cols - 1,
                       data_start, data_end)
        row += 2   # blank separator

    # ── 9A – B2B Regular ──
    sec_title("9A - Amendment to B2B Regular (Table 4)", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Differential",8,12)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess",
             "Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['9A_4A_amt'][2],T['9A_4A_amt'][3],T['9A_4A_amt'][4],
        T['9A_4A_amt'][5],T['9A_4A_amt'][6],
        T['9A_4A_diff'][0],T['9A_4A_diff'][1],T['9A_4A_diff'][2],
        T['9A_4A_diff'][3],T['9A_4A_diff'][4],
    ], 10)

    # ── 9A – B2B Reverse Charge ──
    sec_title("9A - Amendment to B2B Reverse Charge (Table 4)", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Differential",8,12)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess",
             "Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['9A_4B_amt'][2],T['9A_4B_amt'][3],T['9A_4B_amt'][4],
        T['9A_4B_amt'][5],T['9A_4B_amt'][6],
        T['9A_4B_diff'][0],T['9A_4B_diff'][1],T['9A_4B_diff'][2],
        T['9A_4B_diff'][3],T['9A_4B_diff'][4],
    ], 10)

    # ── 9A – B2CL ──
    sec_title("9A - Amendment to B2CL (Table 5)", 8)
    sub_hdr([("",1,2),("Amended Amount",3,5),("Net Differential",6,8)])
    fld_hdr(["","Month","Value","IGST","Cess","Value","IGST","Cess"])
    data_rows(lambda h,T: [
        T['9A_5_amt'][2],T['9A_5_amt'][3],
        T['9A_5_amt'][4] if len(T['9A_5_amt'])>4 else 0,
        T['9A_5_diff'][0],T['9A_5_diff'][1],
        T['9A_5_diff'][2] if len(T['9A_5_diff'])>2 else 0,
    ], 6)

    # ── 9A – 6A Exports ──
    sec_title("9A - Amendment to Exports (Table 6A)", 12)
    sub_hdr([("",1,2),("Amended Total",3,5),("Net Diff Total",6,8),
             ("EXPWP Amt",9,11),("EXPWOP Amt",12,12)])
    fld_hdr(["","Month",
             "Value","IGST","Cess","Value","IGST","Cess",
             "Value","IGST","Cess","Value"])
    data_rows(lambda h,T: [
        T['9A_6A_amt'][2],T['9A_6A_amt'][3],
        T['9A_6A_amt'][4] if len(T['9A_6A_amt'])>4 else 0,
        T['9A_6A_diff'][0],T['9A_6A_diff'][1],
        T['9A_6A_diff'][2] if len(T['9A_6A_diff'])>2 else 0,
        T['9A_6A_EXPWP_amt'][2],T['9A_6A_EXPWP_amt'][3],
        T['9A_6A_EXPWP_amt'][4] if len(T['9A_6A_EXPWP_amt'])>4 else 0,
        T['9A_6A_EXPWOP_amt'][2] if len(T['9A_6A_EXPWOP_amt'])>2 else 0,
    ], 10)

    # ── 9A – 6B SEZ ──
    sec_title("9A - Amendment to SEZ (Table 6B)", 12)
    sub_hdr([("",1,2),("Amended Total",3,5),("Net Diff Total",6,8),
             ("SEZWP Amt",9,11),("SEZWOP Amt",12,12)])
    fld_hdr(["","Month",
             "Value","IGST","Cess","Value","IGST","Cess",
             "Value","IGST","Cess","Value"])
    data_rows(lambda h,T: [
        T['9A_6B_amt'][2],T['9A_6B_amt'][3],
        T['9A_6B_amt'][4] if len(T['9A_6B_amt'])>4 else 0,
        T['9A_6B_diff'][0],T['9A_6B_diff'][1],
        T['9A_6B_diff'][2] if len(T['9A_6B_diff'])>2 else 0,
        T['9A_6B_SEZWP_amt'][2],T['9A_6B_SEZWP_amt'][3],
        T['9A_6B_SEZWP_amt'][4] if len(T['9A_6B_SEZWP_amt'])>4 else 0,
        T['9A_6B_SEZWOP_amt'][2] if len(T['9A_6B_SEZWOP_amt'])>2 else 0,
    ], 10)

    # ── 9A – 6C Deemed Exports ──
    sec_title("9A - Amendment to Deemed Exports (Table 6C)", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Differential",8,12)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess",
             "Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['9A_6C_amt'][2],T['9A_6C_amt'][3],T['9A_6C_amt'][4],
        T['9A_6C_amt'][5],T['9A_6C_amt'][6],
        T['9A_6C_diff'][0],T['9A_6C_diff'][1],T['9A_6C_diff'][2],
        T['9A_6C_diff'][3],T['9A_6C_diff'][4],
    ], 10)

    # ── 9C – CDNRA ──
    sec_title("9C - Amended Credit/Debit Notes (Registered) - CDNRA", 22)
    sub_hdr([("",1,2),("Amended Total",3,7),("Net Diff Total",8,12),
             ("B2B Reg Net Diff",13,17),("B2B RC Net Diff",18,22)])
    fld_hdr(["","Month"]+["Value","IGST","CGST","SGST","Cess"]*4)
    data_rows(lambda h,T: [
        T['9C_CDNRA_amt'][2],T['9C_CDNRA_amt'][3],T['9C_CDNRA_amt'][4],
        T['9C_CDNRA_amt'][5],T['9C_CDNRA_amt'][6],
        T['9C_CDNRA_diff'][0],T['9C_CDNRA_diff'][1],T['9C_CDNRA_diff'][2],
        T['9C_CDNRA_diff'][3],T['9C_CDNRA_diff'][4],
        T['9C_CDNRA_B2B_diff'][2],T['9C_CDNRA_B2B_diff'][3],
        T['9C_CDNRA_B2B_diff'][4],T['9C_CDNRA_B2B_diff'][5],
        T['9C_CDNRA_B2B_diff'][6],
        T['9C_CDNRA_RC_diff'][2],T['9C_CDNRA_RC_diff'][3],
        T['9C_CDNRA_RC_diff'][4],T['9C_CDNRA_RC_diff'][5],
        T['9C_CDNRA_RC_diff'][6],
    ], 20)

    # ── 9C – CDNURA ──
    sec_title("9C - Amended Credit/Debit Notes (Unregistered) - CDNURA", 14)
    sub_hdr([("",1,2),("Amended Total",3,5),("Net Diff Total",6,8),
             ("B2CL",9,11),("EXPWP",12,14)])
    fld_hdr(["","Month",
             "Value","IGST","Cess","Value","IGST","Cess",
             "Value","IGST","Cess","Value","IGST","Cess"])
    data_rows(lambda h,T: [
        T['9C_CDNURA_amt'][2] if len(T['9C_CDNURA_amt'])>2 else 0,
        T['9C_CDNURA_amt'][3] if len(T['9C_CDNURA_amt'])>3 else 0,
        T['9C_CDNURA_amt'][4] if len(T['9C_CDNURA_amt'])>4 else 0,
        T['9C_CDNURA_diff'][0] if len(T['9C_CDNURA_diff'])>0 else 0,
        T['9C_CDNURA_diff'][1] if len(T['9C_CDNURA_diff'])>1 else 0,
        T['9C_CDNURA_diff'][2] if len(T['9C_CDNURA_diff'])>2 else 0,
        T['9C_CDNURA_B2CL'][2] if len(T['9C_CDNURA_B2CL'])>2 else 0,
        T['9C_CDNURA_B2CL'][3] if len(T['9C_CDNURA_B2CL'])>3 else 0,
        T['9C_CDNURA_B2CL'][4] if len(T['9C_CDNURA_B2CL'])>4 else 0,
        T['9C_CDNURA_EXPWP'][2] if len(T['9C_CDNURA_EXPWP'])>2 else 0,
        T['9C_CDNURA_EXPWP'][3] if len(T['9C_CDNURA_EXPWP'])>3 else 0,
        T['9C_CDNURA_EXPWP'][4] if len(T['9C_CDNURA_EXPWP'])>4 else 0,
    ], 12)

    # ── 10 – B2C Amendment ──
    sec_title("10 - Amendment to B2C (Others)", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Differential",8,12)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess",
             "Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['10_amt'][2],T['10_amt'][3],T['10_amt'][4],
        T['10_amt'][5],T['10_amt'][6],
        T['10_diff'][0],T['10_diff'][1],T['10_diff'][2],
        T['10_diff'][3],T['10_diff'][4],
    ], 10)

    # ── 11A(1) ──
    sec_title("11A(1), 11A(2) - Advances Received", 7)
    sub_hdr([("",1,2),("Total",3,7)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['11A1'][2],T['11A1'][3],T['11A1'][4],T['11A1'][5],T['11A1'][6]
    ], 5)

    # ── 11B(1) ──
    sec_title("11B(1), 11B(2) - Advance Adjusted", 7)
    sub_hdr([("",1,2),("Total",3,7)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['11B1'][2],T['11B1'][3],T['11B1'][4],T['11B1'][5],T['11B1'][6]
    ], 5)

    # ── 11A Amendment ──
    sec_title("11A - Amendment to Advances Received", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Total",8,12)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess",
             "Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['11A_amt'][2],T['11A_amt'][3],T['11A_amt'][4],
        T['11A_amt'][5],T['11A_amt'][6],
        T['11A_diff'][0],T['11A_diff'][1],T['11A_diff'][2],
        T['11A_diff'][3],T['11A_diff'][4],
    ], 10)

    # ── 11B Amendment ──
    sec_title("11B - Amendment to Advances Adjusted", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Total",8,12)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess",
             "Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['11B_amt'][2],T['11B_amt'][3],T['11B_amt'][4],
        T['11B_amt'][5],T['11B_amt'][6],
        T['11B_diff'][0],T['11B_diff'][1],T['11B_diff'][2],
        T['11B_diff'][3],T['11B_diff'][4],
    ], 10)

    # ── 12 – HSN ──
    sec_title("12 - HSN-wise Summary of Outward Supplies", 19)
    sub_hdr([("",1,2),("Total",3,8),("B2B Total",9,14),("B2C Total",15,19)])
    fld_hdr(["","Month"]+["No.Rec","Value","IGST","CGST","SGST","Cess"]*3)
    data_rows(lambda h,T: [
        T['12_total'][0],
        T['12_total'][2] if len(T['12_total'])>2 else 0,
        T['12_total'][3] if len(T['12_total'])>3 else 0,
        T['12_total'][4] if len(T['12_total'])>4 else 0,
        T['12_total'][5] if len(T['12_total'])>5 else 0,
        T['12_total'][6] if len(T['12_total'])>6 else 0,
        T['12_B2B'][0],
        T['12_B2B'][2] if len(T['12_B2B'])>2 else 0,
        T['12_B2B'][3] if len(T['12_B2B'])>3 else 0,
        T['12_B2B'][4] if len(T['12_B2B'])>4 else 0,
        T['12_B2B'][5] if len(T['12_B2B'])>5 else 0,
        T['12_B2B'][6] if len(T['12_B2B'])>6 else 0,
        T['12_B2C'][0],
        T['12_B2C'][2] if len(T['12_B2C'])>2 else 0,
        T['12_B2C'][3] if len(T['12_B2C'])>3 else 0,
        T['12_B2C'][4] if len(T['12_B2C'])>4 else 0,
        T['12_B2C'][5] if len(T['12_B2C'])>5 else 0,
        T['12_B2C'][6] if len(T['12_B2C'])>6 else 0,
    ], 18)

    # ── 13 – Documents ──
    sec_title("13 - Documents Issued", 3)
    fld_hdr(["","Month","Net Issued Documents"])
    data_rows(lambda h,T: [T['13']], 1)

    # ── 14 – E-Commerce ──
    sec_title("14 - Supplies through E-Commerce Operators", 17)
    sub_hdr([("",1,2),("Total",3,7),("Liable u/s 52",8,12),
             ("Liable u/s 9(5)",13,17)])
    fld_hdr(["","Month"]+["Value","IGST","CGST","SGST","Cess"]*3)
    data_rows(lambda h,T: [
        T['14_total'][2],T['14_total'][3],T['14_total'][4],
        T['14_total'][5],T['14_total'][6],
        T['14_52'][2],T['14_52'][3],T['14_52'][4],
        T['14_52'][5],T['14_52'][6],
        T['14_95'][2],T['14_95'][3],T['14_95'][4],
        T['14_95'][5],T['14_95'][6],
    ], 15)

    # ── 14A – Amended E-Commerce ──
    sec_title("14A - Amended E-Commerce Supplies", 12)
    sub_hdr([("",1,2),("Amended Total",3,7),("Net Differential",8,12)])
    fld_hdr(["","Month","Value","IGST","CGST","SGST","Cess",
             "Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['14A_amt'][2],T['14A_amt'][3],T['14A_amt'][4],
        T['14A_amt'][5],T['14A_amt'][6],
        T['14A_diff'][2] if len(T['14A_diff'])>2 else 0,
        T['14A_diff'][3] if len(T['14A_diff'])>3 else 0,
        T['14A_diff'][4] if len(T['14A_diff'])>4 else 0,
        T['14A_diff'][5] if len(T['14A_diff'])>5 else 0,
        T['14A_diff'][6] if len(T['14A_diff'])>6 else 0,
    ], 10)

    # ── 15 – Supplies u/s 9(5) ──
    sec_title("15 - Supplies U/s 9(5)", 12)
    sub_hdr([("",1,2),("Total",3,7),("For Registered Recipients",8,12)])
    fld_hdr(["","Month"]+["Value","IGST","CGST","SGST","Cess"]*2)
    data_rows(lambda h,T: [
        T['15_total'][0] if isinstance(T['15_total'][0],float) else 0,
        T['15_total'][1] if len(T['15_total'])>1 else 0,
        T['15_total'][2] if len(T['15_total'])>2 else 0,
        T['15_total'][3] if len(T['15_total'])>3 else 0,
        T['15_total'][4] if len(T['15_total'])>4 else 0,
        T['15_reg'][0] if isinstance(T['15_reg'][0],float) else 0,
        T['15_reg'][1] if len(T['15_reg'])>1 else 0,
        T['15_reg'][2] if len(T['15_reg'])>2 else 0,
        T['15_reg'][3] if len(T['15_reg'])>3 else 0,
        T['15_reg'][4] if len(T['15_reg'])>4 else 0,
    ], 10)

    # ── 15A(I) ──
    sec_title("15A(I) - Amended Supplies U/s 9(5) – For Registered", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Differential",8,12)])
    fld_hdr(["","Month"]+["Value","IGST","CGST","SGST","Cess"]*2)
    data_rows(lambda h,T: [
        T['15A1_amt'][0] if isinstance(T['15A1_amt'][0],float) else 0,
        T['15A1_amt'][1] if len(T['15A1_amt'])>1 else 0,
        T['15A1_amt'][2] if len(T['15A1_amt'])>2 else 0,
        T['15A1_amt'][3] if len(T['15A1_amt'])>3 else 0,
        T['15A1_amt'][4] if len(T['15A1_amt'])>4 else 0,
        T['15A1_diff'][0] if isinstance(T['15A1_diff'][0],float) else 0,
        T['15A1_diff'][1] if len(T['15A1_diff'])>1 else 0,
        T['15A1_diff'][2] if len(T['15A1_diff'])>2 else 0,
        T['15A1_diff'][3] if len(T['15A1_diff'])>3 else 0,
        T['15A1_diff'][4] if len(T['15A1_diff'])>4 else 0,
    ], 10)

    # ── 15A(II) ──
    sec_title("15A(II) - Amended Supplies U/s 9(5) – For Unregistered", 12)
    sub_hdr([("",1,2),("Amended Amount",3,7),("Net Differential",8,12)])
    fld_hdr(["","Month"]+["Value","IGST","CGST","SGST","Cess"]*2)
    data_rows(lambda h,T: [
        T['15A2_amt'][0] if isinstance(T['15A2_amt'][0],float) else 0,
        T['15A2_amt'][1] if len(T['15A2_amt'])>1 else 0,
        T['15A2_amt'][2] if len(T['15A2_amt'])>2 else 0,
        T['15A2_amt'][3] if len(T['15A2_amt'])>3 else 0,
        T['15A2_amt'][4] if len(T['15A2_amt'])>4 else 0,
        T['15A2_diff'][0] if isinstance(T['15A2_diff'][0],float) else 0,
        T['15A2_diff'][1] if len(T['15A2_diff'])>1 else 0,
        T['15A2_diff'][2] if len(T['15A2_diff'])>2 else 0,
        T['15A2_diff'][3] if len(T['15A2_diff'])>3 else 0,
        T['15A2_diff'][4] if len(T['15A2_diff'])>4 else 0,
    ], 10)

    # ── Total Liability ──
    sec_title("Total Liability (Outward supplies other than Reverse charge)", 7)
    fld_hdr(["","Month","Taxable Value","IGST","CGST","SGST","Cess"])
    data_rows(lambda h,T: [
        T['total_liability'][0], T['total_liability'][1],
        T['total_liability'][2], T['total_liability'][3],
        T['total_liability'][4],
    ], 5)



def _build_all_data(ws, all_data):
    """Flat table with every field from every PDF."""
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 55
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 14

    headers = [
        "Financial Year","GSTIN","Legal Name","ARN Date","Tax Period",
        "Table No.","Description","No. of Records","Document Type",
        "Value","Integrated Tax","Central Tax","State Tax","Cess"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        style_cell(c, h, bold=True, bg=C_HEADER_DARK, font_color=C_WHITE, size=10)

    def flat_row(ws, row_num, hdr, T, table_no, desc, no_rec, doc_type,
                 value, igst, cgst, sgst, cess):
        vals = [
            hdr['financial_year'], hdr['gstin'], hdr['legal_name'],
            hdr['arn_date'], hdr['tax_period'].upper(),
            table_no, desc, no_rec, doc_type,
            value, igst, cgst, sgst, cess
        ]
        bg = C_WHITE if row_num % 2 == 0 else "F0F7FF"
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col)
            if col in (10, 11, 12, 13, 14):
                num_cell(c, v, bg=bg)
            else:
                style_cell(c, v, bg=bg,
                           h_align="left" if col in (3, 6, 7) else "center")
        return row_num + 1

    row = 2
    for d in all_data:
        h, T = d['hdr'], d['T']

        row = flat_row(ws,row,h,T,"4A","4A - B2B Regular - Total",
            T['4A'][0],T['4A'][1],T['4A'][2],T['4A'][3],T['4A'][4],T['4A'][5],T['4A'][6])
        row = flat_row(ws,row,h,T,"4B","4B - B2B Reverse Charge - Total",
            T['4B'][0],T['4B'][1],T['4B'][2],T['4B'][3],T['4B'][4],T['4B'][5],T['4B'][6])
        row = flat_row(ws,row,h,T,"5","5 - B2CL (Large) - Total",
            T['5'][0],T['5'][1],T['5'][2],T['5'][3],0,0,
            T['5'][4] if len(T['5'])>4 else 0)
        row = flat_row(ws,row,h,T,"6A","6A - Exports Total",
            T['6A'][0],T['6A'][1],T['6A'][2],T['6A'][3],0,0,
            T['6A'][4] if len(T['6A'])>4 else 0)
        row = flat_row(ws,row,h,T,"6A","6A - EXPWP",
            T['6A_EXPWP'][0],T['6A_EXPWP'][1],T['6A_EXPWP'][2],T['6A_EXPWP'][3],0,0,
            T['6A_EXPWP'][4] if len(T['6A_EXPWP'])>4 else 0)
        row = flat_row(ws,row,h,T,"6A","6A - EXPWOP",
            T['6A_EXPWOP'][0],T['6A_EXPWOP'][1],
            T['6A_EXPWOP'][2] if len(T['6A_EXPWOP'])>2 else 0,0,0,0,0)
        row = flat_row(ws,row,h,T,"6B","6B - SEZ Total",
            T['6B'][0],T['6B'][1],T['6B'][2],T['6B'][3],0,0,
            T['6B'][4] if len(T['6B'])>4 else 0)
        row = flat_row(ws,row,h,T,"6B","6B - SEZWP",
            T['6B_SEZWP'][0],T['6B_SEZWP'][1],T['6B_SEZWP'][2],T['6B_SEZWP'][3],0,0,
            T['6B_SEZWP'][4] if len(T['6B_SEZWP'])>4 else 0)
        row = flat_row(ws,row,h,T,"6B","6B - SEZWOP",
            T['6B_SEZWOP'][0],T['6B_SEZWOP'][1],
            T['6B_SEZWOP'][2] if len(T['6B_SEZWOP'])>2 else 0,0,0,0,0)
        row = flat_row(ws,row,h,T,"6C","6C - Deemed Exports Total",
            T['6C'][0],T['6C'][1],T['6C'][2],T['6C'][3],T['6C'][4],T['6C'][5],T['6C'][6])
        row = flat_row(ws,row,h,T,"7","7 - B2CS (Others) - Total",
            T['7'][0],T['7'][1],T['7'][2],T['7'][3],T['7'][4],T['7'][5],T['7'][6])
        row = flat_row(ws,row,h,T,"8","8 - Nil/Exempted/Non-GST Total",
            0,"NA",T['8'][0],0,0,0,0)
        row = flat_row(ws,row,h,T,"8","8 - Nil",0,"NA",T['8'][1],0,0,0,0)
        row = flat_row(ws,row,h,T,"8","8 - Exempted",0,"NA",T['8'][2],0,0,0,0)
        row = flat_row(ws,row,h,T,"8","8 - Non-GST",0,"NA",T['8'][3],0,0,0,0)
        row = flat_row(ws,row,h,T,"9A","9A - B2B Regular Amended Amt",
            T['9A_4A_amt'][0],T['9A_4A_amt'][1],T['9A_4A_amt'][2],T['9A_4A_amt'][3],
            T['9A_4A_amt'][4],T['9A_4A_amt'][5],T['9A_4A_amt'][6])
        row = flat_row(ws,row,h,T,"9A","9A - B2B Regular Net Diff",
            0,"NA",T['9A_4A_diff'][0],T['9A_4A_diff'][1],T['9A_4A_diff'][2],
            T['9A_4A_diff'][3],T['9A_4A_diff'][4])
        row = flat_row(ws,row,h,T,"9A","9A - B2B RC Amended Amt",
            T['9A_4B_amt'][0],T['9A_4B_amt'][1],T['9A_4B_amt'][2],T['9A_4B_amt'][3],
            T['9A_4B_amt'][4],T['9A_4B_amt'][5],T['9A_4B_amt'][6])
        row = flat_row(ws,row,h,T,"9A","9A - B2B RC Net Diff",
            0,"NA",T['9A_4B_diff'][0],T['9A_4B_diff'][1],T['9A_4B_diff'][2],
            T['9A_4B_diff'][3],T['9A_4B_diff'][4])
        row = flat_row(ws,row,h,T,"9A","9A - B2CL Amended Amt",
            T['9A_5_amt'][0],T['9A_5_amt'][1],T['9A_5_amt'][2],T['9A_5_amt'][3],0,0,
            T['9A_5_amt'][4] if len(T['9A_5_amt'])>4 else 0)
        row = flat_row(ws,row,h,T,"9A","9A - B2CL Net Diff",
            0,"NA",T['9A_5_diff'][0],T['9A_5_diff'][1],0,0,
            T['9A_5_diff'][2] if len(T['9A_5_diff'])>2 else 0)
        row = flat_row(ws,row,h,T,"9B-CDNR","9B CDNR - Total Net off",
            T['9B_CDNR_total'][0],T['9B_CDNR_total'][1],T['9B_CDNR_total'][2],
            T['9B_CDNR_total'][3],T['9B_CDNR_total'][4],T['9B_CDNR_total'][5],
            T['9B_CDNR_total'][6])
        row = flat_row(ws,row,h,T,"9B-CDNR","9B CDNR - B2B Regular Net Total",
            T['9B_CDNR_B2B'][0],T['9B_CDNR_B2B'][1],T['9B_CDNR_B2B'][2],
            T['9B_CDNR_B2B'][3],T['9B_CDNR_B2B'][4],T['9B_CDNR_B2B'][5],
            T['9B_CDNR_B2B'][6])
        row = flat_row(ws,row,h,T,"9B-CDNR","9B CDNR - B2B RC Net Total",
            T['9B_CDNR_RC'][0],T['9B_CDNR_RC'][1],T['9B_CDNR_RC'][2],
            T['9B_CDNR_RC'][3],T['9B_CDNR_RC'][4],T['9B_CDNR_RC'][5],
            T['9B_CDNR_RC'][6])
        row = flat_row(ws,row,h,T,"9B-CDNUR","9B CDNUR - Total Net off",
            T['9B_CDNUR_total'][0],T['9B_CDNUR_total'][1],T['9B_CDNUR_total'][2],
            T['9B_CDNUR_total'][3],0,0,
            T['9B_CDNUR_total'][4] if len(T['9B_CDNUR_total'])>4 else 0)
        row = flat_row(ws,row,h,T,"9B-CDNUR","9B CDNUR - B2CL",
            T['9B_CDNUR_B2CL'][0],T['9B_CDNUR_B2CL'][1],T['9B_CDNUR_B2CL'][2],
            T['9B_CDNUR_B2CL'][3],0,0,
            T['9B_CDNUR_B2CL'][4] if len(T['9B_CDNUR_B2CL'])>4 else 0)
        row = flat_row(ws,row,h,T,"10","10 - B2C Others Amended Amt",
            T['10_amt'][0],T['10_amt'][1],T['10_amt'][2],T['10_amt'][3],
            T['10_amt'][4],T['10_amt'][5],T['10_amt'][6])
        row = flat_row(ws,row,h,T,"10","10 - B2C Others Net Diff",
            0,"NA",T['10_diff'][0],T['10_diff'][1],T['10_diff'][2],
            T['10_diff'][3],T['10_diff'][4])
        row = flat_row(ws,row,h,T,"11A(1)","11A(1),11A(2) - Advances Received",
            T['11A1'][0],T['11A1'][1],T['11A1'][2],T['11A1'][3],
            T['11A1'][4],T['11A1'][5],T['11A1'][6])
        row = flat_row(ws,row,h,T,"11B(1)","11B(1),11B(2) - Advance Adjusted",
            T['11B1'][0],T['11B1'][1],T['11B1'][2],T['11B1'][3],
            T['11B1'][4],T['11B1'][5],T['11B1'][6])
        row = flat_row(ws,row,h,T,"12","12 - HSN Summary Total",
            T['12_total'][0],"NA",
            T['12_total'][2] if len(T['12_total'])>2 else 0,
            T['12_total'][3] if len(T['12_total'])>3 else 0,
            T['12_total'][4] if len(T['12_total'])>4 else 0,
            T['12_total'][5] if len(T['12_total'])>5 else 0,
            T['12_total'][6] if len(T['12_total'])>6 else 0)
        row = flat_row(ws,row,h,T,"12","12 - HSN Summary B2B Total",
            T['12_B2B'][0],"NA",
            T['12_B2B'][2] if len(T['12_B2B'])>2 else 0,
            T['12_B2B'][3] if len(T['12_B2B'])>3 else 0,
            T['12_B2B'][4] if len(T['12_B2B'])>4 else 0,
            T['12_B2B'][5] if len(T['12_B2B'])>5 else 0,
            T['12_B2B'][6] if len(T['12_B2B'])>6 else 0)
        row = flat_row(ws,row,h,T,"12","12 - HSN Summary B2C Total",
            T['12_B2C'][0],"NA",
            T['12_B2C'][2] if len(T['12_B2C'])>2 else 0,
            T['12_B2C'][3] if len(T['12_B2C'])>3 else 0,
            T['12_B2C'][4] if len(T['12_B2C'])>4 else 0,
            T['12_B2C'][5] if len(T['12_B2C'])>5 else 0,
            T['12_B2C'][6] if len(T['12_B2C'])>6 else 0)
        row = flat_row(ws,row,h,T,"13","13 - Documents Issued",
            T['13'],"All Documents",0,0,0,0,0)
        row = flat_row(ws,row,h,T,"14","14 - E-Commerce Total",
            T['14_total'][0],T['14_total'][1],T['14_total'][2],T['14_total'][3],
            T['14_total'][4],T['14_total'][5],T['14_total'][6])
        row = flat_row(ws,row,h,T,"Total","Total Liability",
            0,"NA",T['total_liability'][0],T['total_liability'][1],
            T['total_liability'][2],T['total_liability'][3],T['total_liability'][4])

# STREAMLIT UI
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="GSTR-1 Consolidator",
    page_icon="📊",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .main { background-color: #f8f9fa; }
    .stApp { font-family: Arial, sans-serif; }
    .title-box {
        background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
        padding: 25px 30px;
        border-radius: 12px;
        margin-bottom: 20px;
    }
    .title-box h1 { color: white; margin: 0; font-size: 28px; }
    .title-box p  { color: #BDD7EE; margin: 5px 0 0 0; font-size: 14px; }
    .info-card {
        background: white;
        border: 1px solid #BDD7EE;
        border-left: 4px solid #2E75B6;
        border-radius: 8px;
        padding: 15px 20px;
        margin: 10px 0;
    }
    .success-card {
        background: #E2EFDA;
        border: 1px solid #70AD47;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #1F3864, #2E75B6) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 10px 24px !important;
        font-size: 15px !important;
        font-weight: bold !important;
        width: 100%;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="title-box">
    <h1>📊 GSTR-1 PDF Consolidator</h1>
    <p>Upload multiple GSTR-1 PDFs → Get a consolidated Excel report with all tables</p>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns([2, 1])

with col1:
    st.markdown("### 📁 Upload GSTR-1 PDFs")
    uploaded_files = st.file_uploader(
        "Drag and drop or browse GSTR-1 PDF files",
        type=["pdf"],
        accept_multiple_files=True,
        help="Upload one or more GSTR-1 PDF files. You can mix different GSTINs and financial years."
    )

with col2:
    st.markdown("### ℹ️ How it works")
    st.markdown("""
    <div class="info-card">
    <b>Steps:</b><br>
    1. Upload your GSTR-1 PDFs<br>
    2. Preview parsed data<br>
    3. Click Generate Excel<br>
    4. Download the report<br><br>
    <b>Output Sheets:</b><br>
    • SUMMARY<br>
    • sales, exports, sez<br>
    • CDNR<br>
    • Remaining (9A–15A)<br>
    • all data (flat table)
    </div>
    """, unsafe_allow_html=True)

if uploaded_files:
    st.markdown("---")
    st.markdown(f"### 📋 Processing {len(uploaded_files)} file(s)")

    all_data = []
    errors = []

    progress = st.progress(0)
    status = st.empty()

    for i, f in enumerate(uploaded_files):
        status.text(f"Parsing: {f.name}...")
        try:
            hdr, T = parse_gstr1_pdf(f.read())
            all_data.append({'hdr': hdr, 'T': T, 'filename': f.name})
        except Exception as e:
            errors.append(f"{f.name}: {str(e)}")
        progress.progress((i + 1) / len(uploaded_files))

    status.empty()
    progress.empty()

    if errors:
        st.error("⚠️ Errors parsing some files:\n" + "\n".join(errors))

    if all_data:
        # Preview table
        st.markdown("### ✅ Parsed Files")
        preview_rows = []
        for d in all_data:
            h = d['hdr']
            preview_rows.append({
                "File": d['filename'],
                "GSTIN": h['gstin'],
                "Legal Name": h['legal_name'],
                "FY": h['financial_year'],
                "Period": h['tax_period'],
                "ARN": h['arn'],
                "ARN Date": h['arn_date'],
                "4A B2B Value": f"₹{d['T']['4A'][2]:,.2f}",
                "9B CDNR Value": f"₹{d['T']['9B_CDNR_total'][2]:,.2f}",
                "Total Liability Value": f"₹{d['T']['total_liability'][0]:,.2f}",
            })
        import pandas as pd
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True)

        st.markdown("---")
        col_gen, col_dl = st.columns([1, 1])

        with col_gen:
            if st.button("⚡ Generate Consolidated Excel", use_container_width=True):
                with st.spinner("Building Excel workbook..."):
                    excel_buf = build_excel(all_data)
                    fname = f"GSTR1_Consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    st.session_state['excel_buf'] = excel_buf.getvalue()
                    st.session_state['excel_fname'] = fname
                st.success("✅ Excel generated successfully!")

        if 'excel_buf' in st.session_state:
            with col_dl:
                st.download_button(
                    label="⬇️ Download Excel Report",
                    data=st.session_state['excel_buf'],
                    file_name=st.session_state['excel_fname'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            st.markdown(f"""
            <div class="success-card">
            ✅ <b>Ready:</b> {st.session_state['excel_fname']}<br>
            📑 5 sheets: SUMMARY · sales,exports,sez · CDNR · Remaining · all data
            </div>
            """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div class="info-card" style="text-align:center; padding: 40px;">
    <h3 style="color:#2E75B6;">👆 Upload GSTR-1 PDFs to get started</h3>
    <p style="color:#666;">Supports multiple PDFs — mix different GSTINs, months, and financial years</p>
    </div>
    """, unsafe_allow_html=True)
